"""012 · GraphRAG 领域建图 · T2 簇(CSV/表格结构化入图 → custom_kg 直灌，不过 LLM)。

CSV→KG 映射契约：

- CSV/表格的行列语义**本就确定**：一行 = 一条记录、列 = 角色(实体/属性/关系)。故走
  LightRAG `ainsert_custom_kg` **结构化直灌**(行→实体、列关系→边、行合成 chunk)，
  **不调 LLM 抽取**(`llm_model_func` 零调用)、实体名与单元格**逐字一致**(无 LLM 改写)。
  区别于"把 CSV 当文本喂 LLM 乱抽"(AM-710)。
- **列类型推断契约**(FR-327 / AM-708)：按表头关键词把每列判为
  实体列(entity)/ 属性列(attribute)/ 关系列(relation)，各带 (role, confidence)。
- **N 行批量 + file_path 锚点**(FR-328/329/336 / AM-709)：N 行 → N 个主体实体，
  每实体带 `file_path` 锚点指向该表(贯通 011 检索侧 references)。
- **脏数据韧性 · 部分成功可见**(FR-331 / AM-711)：空列跳过、malformed 行进诊断报告，
  **不整表失败、不静默丢数据**、中英混合正常处理。
- **dry-run 预览**(FR-330 / AM-712)：预览"将建 N 实体/M 关系 + 类型分布 + 样例"且不写图；
  apply 实际计数 ≡ 预览计数。
- **非 UTF-8 诚实拒收**(FR-331 / AM-715)：非 UTF-8 字节 → GraphRagError，不乱码入图。

映射的决策真身在此 Python 引擎(detect/路由在 documents.py，薄透传在 platform-store)。
"""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field

from graph_rag.core.errors import GraphRagError

_NO_INGEST_LEASE = object()

# ── 列角色推断关键词(SSOT，防魔法串散落)──────────────────────────────────────────────
#
# 顺序敏感：person → supplier → customer → relation → 通用名 → attribute。
# 关系列关键词刻意**不含裸「所属」**——「所属行业」是属性(行业)、「所属集团」才是关系(集团)。

_PERSON_KEYS = ("对接人", "联系人", "负责人", "经办人", "接口人", "联系", "对接", "contact")
_SUPPLIER_KEYS = ("供应商", "供货商", "供货", "supplier", "vendor")
_CUSTOMER_KEYS = ("客户", "customer", "client")
# 关系列 = 指向另一实体的外键(集团/母公司/上级单位…)，非普通类别属性。
_RELATION_KEYS = ("集团", "母公司", "总公司", "上级", "隶属", "关联方", "所属公司", "所属单位", "所属机构", "所属集团", "group")
# 通用实体名列(既非客户也非供应商也非人员，落「组织」)。
_ORG_NAME_KEYS = ("公司名", "单位名", "机构名", "企业名", "组织名")
_NAME_SUFFIXES = ("名称", "名")

ROLE_ENTITY = "entity"
ROLE_ATTRIBUTE = "attribute"
ROLE_RELATION = "relation"
ROLE_SKIP = "skip"

TYPE_CUSTOMER = "客户"
TYPE_SUPPLIER = "供应商"
TYPE_PERSON = "人员"
TYPE_ORG = "组织"


@dataclass(frozen=True)
class ColumnRole:
    column: str
    index: int
    role: str
    confidence: float
    entity_type: str | None = None
    reason: str | None = None


def _contains_any(haystack: str, needles: tuple[str, ...]) -> bool:
    low = haystack.lower()
    return any(needle in haystack or needle in low for needle in needles)


def infer_column_roles(headers: list[str]) -> dict[str, ColumnRole]:
    """FR-327 / AM-708:逐列推断角色 (role, confidence)。

    - 客户名称 → 实体列(客户)、对接人 → 实体列(人员)、注册资本/所属行业 → 属性列、
      所属集团 → 关系列(外键指向另一实体)。
    - 空/空白表头 → skip(不建实体、不建属性)。
    """
    roles: dict[str, ColumnRole] = {}
    for index, raw in enumerate(headers):
        header = (raw or "").strip()
        if not header:
            roles[raw] = ColumnRole(
                column=raw, index=index, role=ROLE_SKIP, confidence=0.0,
                reason="empty_or_blank_header",
            )
            continue
        if _contains_any(header, _PERSON_KEYS):
            role = ColumnRole(header, index, ROLE_ENTITY, 0.9, entity_type=TYPE_PERSON)
        elif _contains_any(header, _SUPPLIER_KEYS):
            role = ColumnRole(header, index, ROLE_ENTITY, 0.9, entity_type=TYPE_SUPPLIER)
        elif _contains_any(header, _CUSTOMER_KEYS):
            role = ColumnRole(header, index, ROLE_ENTITY, 0.95, entity_type=TYPE_CUSTOMER)
        elif _contains_any(header, _RELATION_KEYS):
            role = ColumnRole(header, index, ROLE_RELATION, 0.85, entity_type=TYPE_ORG)
        elif _contains_any(header, _ORG_NAME_KEYS) or any(header.endswith(sfx) for sfx in _NAME_SUFFIXES):
            role = ColumnRole(header, index, ROLE_ENTITY, 0.7, entity_type=TYPE_ORG)
        else:
            role = ColumnRole(header, index, ROLE_ATTRIBUTE, 0.6)
        roles[header] = role
    return roles


# ── CSV / xlsx 解析(非 UTF-8 诚实拒收;xlsx 与 CSV 走同一 rows→records 路径)──────────────

_RESTKEY = "__extra_fields__"

# xlsx = zip 容器,魔法字节 PK\x03\x04(真实 CSV 文本不会以此二进制头开始)。
_XLSX_MAGIC = b"PK\x03\x04"


def _rows_to_table(rows_raw: list[list[str]]) -> tuple[list[str], list[dict]]:
    """把二维单元格矩阵(rows_raw[0]=表头)规范化为 (headers, records)。

    CSV 与 xlsx 共用此路径 → **同映射对称**(AM-713:xlsx 与等价 CSV 产出相同实体/关系)。
    malformed 行(字段数多于表头)经 restkey 收集,交由 build 阶段判 malformed(不在此丢)。
    """
    if not rows_raw:
        raise GraphRagError("表格内容为空(无表头)", "empty_table")
    headers = [(h or "").strip() for h in rows_raw[0]]
    if not any(headers):
        raise GraphRagError("表格缺少有效表头", "empty_table")
    records: list[dict] = []
    for values in rows_raw[1:]:
        record: dict[str, object] = {}
        for i, header in enumerate(headers):
            record[header if header else f"__col{i}__"] = values[i] if i < len(values) else ""
        if len(values) > len(headers):
            record[_RESTKEY] = values[len(headers):]
        record["__ncols__"] = len(values)
        records.append(record)
    return headers, records


def _cell_to_str(value) -> str:
    """xlsx 单元格 → 字符串(与 CSV 文本单元格对称)。

    整数值浮点(1000.0)规范回 "1000",避免 xlsx 数字列与等价 CSV 文本列不对称。
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _parse_xlsx_bytes(content: bytes) -> tuple[list[str], list[dict]]:
    """openpyxl 读 xlsx 首个非空 sheet → 规范化行矩阵 → 走 CSV 同 rows→records 路径(AM-713)。"""
    try:
        from openpyxl import load_workbook  # type: ignore[import-untyped]
    except ImportError as exc:  # pragma: no cover - 依赖已声明,防御性
        raise GraphRagError("xlsx 解析需要 openpyxl 依赖", "unsupported_file_type") from exc
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - openpyxl 抛型繁多,统一转诚实拒收
        raise GraphRagError(f"xlsx 解析失败：{exc}", "unsupported_file_type") from exc
    try:
        rows_raw: list[list[str]] = []
        chosen = None
        for sheet in workbook.worksheets:
            candidate: list[list[str]] = []
            for row in sheet.iter_rows(values_only=True):
                cells = [_cell_to_str(v) for v in row]
                if any(c.strip() for c in cells):  # 跳过整行空(xlsx 常见拖尾空行)
                    candidate.append(cells)
            if candidate:  # 首个含数据的 sheet
                chosen = candidate
                break
        rows_raw = chosen or []
    finally:
        workbook.close()
    if not rows_raw:
        raise GraphRagError("xlsx 无非空 sheet(无表头)", "empty_table")
    return _rows_to_table(rows_raw)


def parse_table_bytes(content: bytes | str) -> tuple[list[str], list[dict]]:
    """把 CSV/xlsx 字节或 CSV 字符串解析为 (headers, records)。

    - xlsx 字节(PK\\x03\\x04 头)→ openpyxl 读首个非空 sheet(AM-713,与 CSV 对称)。
    - 非 UTF-8 CSV 字节 → GraphRagError(unsupported_file_type),不静默乱码(AM-715)。
    - 空表/无表头 → GraphRagError。
    """
    if isinstance(content, bytes) and content[:4] == _XLSX_MAGIC:
        return _parse_xlsx_bytes(content)
    if isinstance(content, bytes):
        try:
            text = content.decode("utf-8")
        except UnicodeDecodeError as exc:
            raise GraphRagError("CSV 必须是 UTF-8 编码文本(检测到非 UTF-8 字节)", "unsupported_file_type") from exc
    else:
        text = content
    text = text.lstrip("﻿")  # 剥 UTF-8 BOM
    reader = csv.reader(io.StringIO(text))
    try:
        rows_raw = list(reader)
    except csv.Error as exc:
        raise GraphRagError(f"CSV 解析失败：{exc}", "unsupported_file_type") from exc
    return _rows_to_table(rows_raw)


# ── custom_kg 构造(行→实体/关系/chunk)+ 部分成功诊断 ────────────────────────────────


@dataclass
class BuildResult:
    custom_kg: dict
    entity_count: int
    relation_count: int
    type_distribution: dict[str, int]
    samples: list[dict]
    diagnostics: list[dict]
    skipped_columns: list[dict]
    success_rows: int
    failed_rows: int
    data_rows: int
    entity_names: list[str] = field(default_factory=list)

    def summary(self) -> dict:
        return {
            "entity_count": self.entity_count,
            "relation_count": self.relation_count,
            "type_distribution": self.type_distribution,
            "data_rows": self.data_rows,
            "success_rows": self.success_rows,
            "failed_rows": self.failed_rows,
            "skipped_columns": self.skipped_columns,
            "diagnostics": self.diagnostics,
            "samples": self.samples,
        }


def _row_text(headers: list[str], record: dict) -> str:
    parts = []
    for h in headers:
        if not h:
            continue
        val = str(record.get(h, "") or "").strip()
        if val:
            parts.append(f"{h}: {val}")
    return "；".join(parts) if parts else "(空行)"


def build_custom_kg(headers: list[str], records: list[dict], *, filename: str, doc_id: str) -> BuildResult:
    """把行列映射为 LightRAG `custom_kg`(entities/relationships/chunks)+ 诊断报告。

    - 主体实体列 = 第一个非人员实体列(客户/供应商/组织)；无则第一个实体列。
    - 每行：主体成实体、人员列成人员实体 + (主体)-[对接]->(人员)、关系列成组织实体 + (主体)-[所属]->(组织)。
    - 空列(整列空)/空白表头列 → skipped_columns(不建空实体)。
    - malformed 行(缺主体值 / 字段数溢出)→ diagnostics(可见)，不整表失败、不静默丢。
    """
    roles = infer_column_roles(headers)
    valid_headers = [h for h in headers if h]

    entity_cols = [h for h in valid_headers if roles[h].role == ROLE_ENTITY]
    person_cols = [h for h in entity_cols if roles[h].entity_type == TYPE_PERSON]
    org_entity_cols = [h for h in entity_cols if roles[h].entity_type != TYPE_PERSON]
    relation_cols = [h for h in valid_headers if roles[h].role == ROLE_RELATION]
    attribute_cols = [h for h in valid_headers if roles[h].role == ROLE_ATTRIBUTE]

    primary_cols = org_entity_cols or entity_cols
    if not primary_cols:
        raise GraphRagError("无法识别任何实体列，无法结构化入图", "no_entity_column")
    primary_col = primary_cols[0]
    primary_type = roles[primary_col].entity_type or TYPE_ORG

    skipped_columns: list[dict] = []
    # 空白表头列 → skip。
    for raw in headers:
        header = (raw or "").strip()
        if not header:
            skipped_columns.append({"column": raw, "reason": "blank_column_name"})
    # 整列全空 → skip(不建空属性/空实体)。
    for h in valid_headers:
        if all(not str(rec.get(h, "") or "").strip() for rec in records):
            skipped_columns.append({"column": h, "reason": "entirely_empty_column"})
    empty_cols = {sc["column"] for sc in skipped_columns if sc["reason"] == "entirely_empty_column"}

    chunks: list[dict] = []
    entities: list[dict] = []
    relationships: list[dict] = []
    diagnostics: list[dict] = []
    success_rows = 0

    def add_entity(name: str, etype: str, source_id: str, description: str) -> None:
        entities.append({
            "entity_name": name,
            "entity_type": etype,
            "description": description or f"{etype}:{name}",
            "source_id": source_id,
            "file_path": filename,
        })

    for i, record in enumerate(records):
        row_no = i + 1  # 1-based 数据行号(不含表头)
        primary_val = str(record.get(primary_col, "") or "").strip()
        # malformed 判定：缺主体值 或 字段数多于表头。
        if not primary_val:
            diagnostics.append({
                "row": row_no, "status": "failed", "reason": "missing_primary_entity_value",
                "column": primary_col,
            })
            continue
        if record.get(_RESTKEY):
            diagnostics.append({
                "row": row_no, "status": "failed",
                "reason": "too_many_fields", "expected": len(headers), "got": record.get("__ncols__"),
            })
            continue

        source_id = f"{doc_id}#row{row_no}"
        chunks.append({
            "content": _row_text(headers, record),
            "source_id": source_id,
            "file_path": filename,
            "chunk_order_index": i,
        })

        # 主体实体描述由属性列合成(跳过空列)。
        attr_bits = []
        for a in attribute_cols:
            if a in empty_cols:
                continue
            v = str(record.get(a, "") or "").strip()
            if v:
                attr_bits.append(f"{a}={v}")
        add_entity(primary_val, primary_type, source_id, "；".join(attr_bits))

        # 人员列 → 人员实体 + (主体)-[对接]->(人员)。
        for pc in person_cols:
            pv = str(record.get(pc, "") or "").strip()
            if not pv:
                continue
            add_entity(pv, TYPE_PERSON, source_id, f"{primary_val} 的对接人")
            relationships.append({
                "src_id": primary_val, "tgt_id": pv,
                "description": f"{primary_val} 对接 {pv}", "keywords": "对接,联系人",
                "weight": 1.0, "source_id": source_id, "file_path": filename,
            })

        # 关系列 → 组织实体 + (主体)-[所属]->(组织)。
        for rc in relation_cols:
            rv = str(record.get(rc, "") or "").strip()
            if not rv:
                continue
            add_entity(rv, TYPE_ORG, source_id, f"{rc}")
            relationships.append({
                "src_id": primary_val, "tgt_id": rv,
                "description": f"{primary_val} 所属 {rv}", "keywords": "所属,隶属",
                "weight": 1.0, "source_id": source_id, "file_path": filename,
            })
        success_rows += 1

    # 去重计数(与 LightRAG ainsert_custom_kg 语义一致：实体按 entity_name、关系按 sorted 端点对)。
    distinct_names: dict[str, str] = {}
    for e in entities:
        distinct_names[e["entity_name"]] = e["entity_type"]
    distinct_pairs = {tuple(sorted((r["src_id"], r["tgt_id"]))) for r in relationships}
    type_distribution: dict[str, int] = {}
    for etype in distinct_names.values():
        type_distribution[etype] = type_distribution.get(etype, 0) + 1

    samples = [
        {"entity_name": name, "entity_type": etype}
        for name, etype in list(distinct_names.items())[:5]
    ]

    custom_kg = {"chunks": chunks, "entities": entities, "relationships": relationships}
    return BuildResult(
        custom_kg=custom_kg,
        entity_count=len(distinct_names),
        relation_count=len(distinct_pairs),
        type_distribution=type_distribution,
        samples=samples,
        diagnostics=diagnostics,
        skipped_columns=skipped_columns,
        success_rows=success_rows,
        failed_rows=len(diagnostics),
        data_rows=len(records),
        entity_names=list(distinct_names.keys()),
    )


# ── dry-run / apply ─────────────────────────────────────────────────────────────────


def preview_structured_ingest(content: bytes | str, *, filename: str, doc_id: str = "preview") -> dict:
    """FR-330 / AM-712:dry-run 预览——构造 custom_kg 但**不写图**，返回预测计数/类型分布/样例。"""
    headers, records = parse_table_bytes(content)
    build = build_custom_kg(headers, records, filename=filename, doc_id=doc_id)
    result = build.summary()
    result["mode"] = "dry_run"
    result["wrote_graph"] = False
    return result


async def ingest_structured_document(
    source,
    doc_id: str,
    content: bytes | str,
    filename: str,
    *,
    batch_size: int = 1000,
    _lease=None,
) -> dict:
    """FR-329 / AM-709/710:CSV/表格结构化入图——构造 custom_kg 走 `ainsert_custom_kg` 直灌。

    - **不过 LLM**:全程只调 `ainsert_custom_kg`，不触 `ainsert(text)`/`llm_model_func` 抽取。
    - N 行批量、每实体带 file_path 锚点、部分成功诊断随返回可见。
    - batch_size:按数据行分批构造多次 custom_kg(大表分批，AM-714 用；小表 = 单批)。
    """
    from graph_rag.core import lightrag_service
    from graph_rag.core.source_operation import source_graph_operation

    if _lease is None:
        async with source_graph_operation(source):
            return await ingest_structured_document(
                source,
                doc_id,
                content,
                filename,
                batch_size=batch_size,
                _lease=_NO_INGEST_LEASE,
            )

    headers, records = parse_table_bytes(content)

    rag = (
        await lightrag_service.get_lightrag_for_ingest(source, doc_id, _lease)
        if _lease is not _NO_INGEST_LEASE
        else await lightrag_service.get_lightrag(source)
    )

    # 按行分批：每批独立构造 custom_kg 调一次 ainsert_custom_kg(合并诊断/计数)。
    batches = [records[i:i + batch_size] for i in range(0, len(records), batch_size)] or [[]]
    agg_names: dict[str, str] = {}
    agg_pairs: set = set()
    diagnostics: list[dict] = []
    skipped_columns: list[dict] = []
    success_rows = 0
    ainsert_calls = 0

    for bi, batch in enumerate(batches):
        build = build_custom_kg(headers, batch, filename=filename, doc_id=f"{doc_id}#b{bi}")
        if build.custom_kg["entities"]:
            await rag.ainsert_custom_kg(build.custom_kg)
            ainsert_calls += 1
        for e in build.custom_kg["entities"]:
            agg_names[e["entity_name"]] = e["entity_type"]
        for r in build.custom_kg["relationships"]:
            agg_pairs.add(tuple(sorted((r["src_id"], r["tgt_id"]))))
        # batch 行号偏移修正到全表 1-based。
        offset = bi * batch_size
        for d in build.diagnostics:
            d = dict(d)
            d["row"] = d["row"] + offset
            diagnostics.append(d)
        for sc in build.skipped_columns:
            if sc not in skipped_columns:
                skipped_columns.append(sc)
        success_rows += build.success_rows

    type_distribution: dict[str, int] = {}
    for etype in agg_names.values():
        type_distribution[etype] = type_distribution.get(etype, 0) + 1

    summary = {
        "entity_count": len(agg_names),
        "relation_count": len(agg_pairs),
        "type_distribution": type_distribution,
        "data_rows": len(records),
        "success_rows": success_rows,
        "failed_rows": len(diagnostics),
        "skipped_columns": skipped_columns,
        "diagnostics": diagnostics,
        "ainsert_custom_kg_calls": ainsert_calls,
        "batches": len(batches),
    }
    return {
        "mode": "apply",
        "wrote_graph": True,
        "entity_names": list(agg_names.keys()),
        "summary": summary,
        **summary,
    }
